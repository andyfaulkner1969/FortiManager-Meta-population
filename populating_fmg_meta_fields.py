import openpyxl
import string
import requests
import json
import urllib3


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

print("***  The device name must match exactly what is in row 1 of the spreadsheet. ***")
print()
device_name = input("Please enter the name of the device: ")

fmg_ip = "xxx.xxx.xxx.xxx"   ##### YOU HAVE TO PUT YOUR FMG IP
user = "api-user"           ##### YOU HAVE TO PUT YOUR USER HERE
password = "password"     #### YOU HAVE TO PUT YOUR password here
filename = '/the/location/of/your/excel/file'   ####  Location of your excel file

#### Excel file needs to be created.  Row 1 will contain the device name (also hostname)
#### The first column is the Meta field names.  They must match Exact.
#### The following columns will have your Meta vlaues.  

url_base="https://" + fmg_ip + "/jsonrpc"

client = requests.session()
#Login request
payload = {
	"id": 1,
	"method": "exec",
	"params": [
		{
			"data": {
				"passwd": password, 
				"user": user     
			},
			"url": "/sys/login/user"
		}
	]
}
r = client.post(url_base, json=payload, verify=False )
#Retrieve session id. Add to HTTP header for future messages parsed_json = json.loads(r.text)
parsed_json = json.loads(r.text)
#print(r.text)
sid = parsed_json['session']
headers = {'session' : sid }

alpha = list(string.ascii_uppercase)    # Supports A-Z columns

wb = openpyxl.load_workbook(filename,data_only=True)
worksheet_names = wb.sheetnames  	#Get the list of sheetnames
sheet_index = worksheet_names.index("Sheet1")  
wb.active = sheet_index  #Set the sheet name to the active sheet
ws = wb.active 	         #Set the  active sheet to the variable

### Here we use the devcie choice and look for a match in the spreedsheet row 1
### This must be an exact match and it is case sensitive

clm_cnt =1
clm_choice = 0
row_found = 0
loop_stop = 0

while loop_stop == 0:  # How we test the loop
	if ws.cell(row=1, column=clm_cnt).value is None: # When none is found stop
		print("Device not found in spreadsheet")
		exit()
		
	elif ws.cell(row=1, column=clm_cnt).value == device_name:
		clm_choice = clm_cnt
		#print(clm_choice) DEBUG
		loop_stop = loop_stop + 1
	else:
		clm_cnt = clm_cnt + 1
		
####  This function will go to the FMG and populate the Meta field to the values it has found 
####  matching the same field.  The field has to match EXACT also.

def populate_meta(meta_name,meta_value):
	global sid
	meta_fields = {
		"id": 1,
		"jsonrpc": "1.0",
		"method": "set",
		"params": [
			{
				"data": {
					"meta fields": {
					meta_name: meta_value,
					
					}
				},
								
					"url": "/dvmdb/device/" + device_name 
			}
		],
		"session": sid,
		"verbose": 1
		
	}

	
	r = client.post(url_base, headers=headers, json=meta_fields, verify=False)
	#print(r.text)  ### DEBUG
	parsed_json = json.loads(r.text)

#	print(parsed_json)  #### DEBUG
	
####  Here is where we go row by row though the spreadsheet to get the field and value calling
####  the function populat_meta

meta_name = ""
meta_value = ""
row_cnt =1
row_found = 0
row_count_stop = 0

while row_count_stop == 0:  # How we test the loop
	if ws.cell(row=row_cnt, column=1).value is None: # When none is found stop
		row_count_stop = row_count_stop + 1
	else:
		meta_name = ws.cell(row=row_cnt, column=1).value
		meta_value = ws.cell(row=row_cnt, column=clm_choice).value
		print("name: ",meta_name, "    value :", meta_value)
		#print(meta_value)
		populate_meta(meta_name, meta_value)
		row_found = row_cnt		
		row_cnt = row_cnt + 1


	




#x = find_row("FGT-C") 
##print(x)  # DEBUG
#
#metadata = {}
#for column in alpha:
#	metadata[ws[column + "1"].value] = ws[column + str(x)].value
#metadata.pop(None)
#		
#data_vlan = metadata["DATA VLAN"]
#data_subnet = metadata["DATA VLAN SUBNET"]
#data_dhcp_s = metadata["DATA DHCP START"]
#data_dhcp_e = metadata["DATA DHCP END"]
#
#print(data_vlan)
#print(data_subnet)
#print(data_dhcp_s)
#print(data_dhcp_e)