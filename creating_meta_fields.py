import requests
import json
import urllib3
import openpyxl
import string

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

fmg_ip = "xxx.xxx.xxx.xxx"   ##### YOU HAVE TO PUT YOUR FMG IP
user = "api-user"           ##### YOU HAVE TO PUT YOUR USER HERE
password = "password"     #### YOU HAVE TO PUT YOUR password here
filename = '/the/location/of/your/excel/file'   ####  Location of your excel file

#### Excel file needs to be created.  Row 1 will contain the device name (also hostname)
#### The first column is the Meta field names.  They must match Exact.
#### The following columns will have your Meta vlaues.  

url_base="https://" + fmg_ip + "/jsonrpc"

client = requests.session()
#Login request
payload = {
	"id": 1,
	"method": "exec",
	"params": [
		{
			"data": {
				"passwd": password,  #### YOU HAVE TO PUT YOUR password here
				"user": user     ##### YOU HAVE TO PUT YOUR USER HERE
			},
			"url": "/sys/login/user"
		}
	]
}
r = client.post(url_base, json=payload, verify=False )
#Retrieve session id. Add to HTTP header for future messages parsed_json = json.loads(r.text)
parsed_json = json.loads(r.text)
#print(r.text)   #### DEBUG
sid = parsed_json['session']
headers = {'session' : sid }
#print(sid)   #### DEBUG


alpha = list(string.ascii_uppercase)    # Supports A-Z columns
wb = openpyxl.load_workbook(filename,data_only=True)
worksheet_names = wb.sheetnames  	#Get the list of sheetnames
#print(worksheet_names)   #### DEBUG
sheet_index = worksheet_names.index("Sheet1")  
#print(sheet_index)   #### DEBUG

wb.active = sheet_index  #Set the sheet name to the active sheet
ws = wb.active 	         #Set the  active sheet to the variable
#print(ws)   #### DEBUG

def find_meta_value():
	
	row_cnt =1
	row_found = 0
	row_count_stop = 0
	meta_dic = []
	
	while row_count_stop == 0:  # How we test the loop
		if ws.cell(row=row_cnt, column=1).value is None: # When none is found stop
			row_count_stop = row_count_stop + 1
		else:
			#print(ws.cell(row=row_cnt, column=1).value) # DEBUG
			meta_dic.append(ws.cell(row=row_cnt, column=1).value)
			row_cnt = row_cnt + 1
	return meta_dic

meta_list = find_meta_value()
#print(meta_list)  # DEBUG

def create_meta(meta_field):
	global sid
	meta_fields = {
		"method": "add",
		"params": [
			{
				"data": [
					{
						"importance": "optional",
						"length": 32,
						"name": meta_field,
						"status": "enable"
					}
				],
				"url": "/dvmdb/_meta_fields/device"
			}
		],
		"session": sid,
		"id": 1
	}

	url_script_list = "https://" + fmg_ip + "/jsonrpc"
	r = client.post(url_base, headers=headers, json=meta_fields, verify=False)
	#print(r.text)
	parsed_json = json.loads(r.text)

	
#	print(parsed_json)
	
meta_field = ""

for x in meta_list:
	print("Meta Field Name :", x)
	meta_field = x
	create_meta(meta_field)